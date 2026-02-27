#!/usr/bin/env python3
"""
NexUS × Clockify — Informe de horas por sprint
Uso: python nexus_clockify.py --api-key TU_API_KEY [--workspace WS_ID] [--sprint 1]
"""

import argparse
import json
import sys
from datetime import datetime, timezone
from pathlib import Path

import requests
import openpyxl
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                               GradientFill)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.series import DataPoint

# ── CONFIG ─────────────────────────────────────────────────────
SIZE_H   = {"XS": 2, "S": 4, "M": 8, "L": 16, "XL": 24}
COLORS   = {
    "header_bg":  "1E1E2E",
    "s1":         "818CF8",
    "s2":         "34D399",
    "s3":         "FBBF24",
    "ok":         "22C55E",
    "warn":       "F59E0B",
    "danger":     "EF4444",
    "light_row":  "F8FAFC",
    "mid_row":    "F1F5F9",
    "white":      "FFFFFF",
    "border":     "E2E8F0",
}
BASE_URL = "https://api.clockify.me/api/v1"


# ── API CLIENT ─────────────────────────────────────────────────
class ClockifyClient:
    def __init__(self, api_key: str):
        self.headers = {"X-Api-Key": api_key, "Content-Type": "application/json"}

    def get(self, path: str, **params) -> dict | list:
        r = requests.get(f"{BASE_URL}{path}", headers=self.headers, params=params)
        r.raise_for_status()
        return r.json()

    def workspace(self) -> str:
        ws = self.get("/workspaces")
        if not ws:
            raise ValueError("No workspaces found for this API key")
        return ws[0]["id"]

    def users(self, ws_id: str) -> dict:
        users = self.get(f"/workspaces/{ws_id}/users", limit=200)
        return {u["id"]: u["name"] for u in users}

    def tags(self, ws_id: str) -> dict:
        tags = self.get(f"/workspaces/{ws_id}/tags", limit=500)
        return {t["id"]: t["name"] for t in tags}

    def entries(self, ws_id: str, start: str | None = None, end: str | None = None) -> list:
        all_entries, page = [], 1
        while True:
            params = {"page": page, "page-size": 200}
            if start: params["start"] = start
            if end:   params["end"]   = end
            batch = self.get(f"/workspaces/{ws_id}/user/CURRENT/time-entries", **params)
            # Use reports API for all users
            break
        # Use reports endpoint for all-users data
        payload = {
            "dateRangeStart": start or "2025-01-01T00:00:00Z",
            "dateRangeEnd":   end   or datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
            "summaryFilter": {"groups": ["USER", "TAG"]},
            "exportType": "JSON"
        }
        r = requests.post(
            f"https://reports.api.clockify.me/v1/workspaces/{ws_id}/reports/detailed",
            headers=self.headers,
            json={
                "dateRangeStart": payload["dateRangeStart"],
                "dateRangeEnd":   payload["dateRangeEnd"],
                "detailedFilter": {"page": 1, "pageSize": 1000, "sortOrder": "DESCENDING"},
            }
        )
        r.raise_for_status()
        data = r.json()
        return data.get("timeentries", [])


# ── BACKLOG LOADER ─────────────────────────────────────────────
def load_backlog(path: str = "nexus-backlog.json") -> dict:
    """Returns {task_id: {title, sprint, area, size, estimated_h}}"""
    p = Path(path)
    if not p.exists():
        # fallback: look in outputs dir
        p = Path(__file__).parent / "nexus-backlog.json"
    with open(p, encoding="utf-8") as f:
        raw = json.load(f)
    tasks = {}
    for sprint_block in raw:
        sprint = sprint_block["sprint"]
        for module in sprint_block["modulos"]:
            for task in module["tareas"]:
                tasks[task["id"]] = {
                    "title":       task["titulo"],
                    "sprint":      sprint,
                    "area":        module["modulo"],
                    "size":        task["talla"],
                    "estimated_h": SIZE_H.get(task["talla"], 0),
                }
    return tasks


# ── DATA PROCESSOR ─────────────────────────────────────────────
def process_entries(entries: list, tags_map: dict, users_map: dict, tasks: dict) -> dict:
    """
    Returns:
      by_task: {task_id: {estimated_h, real_h, by_user: {name: h}}}
      by_user: {name: {total_h, by_task: {task_id: h}}}
    """
    by_task = {tid: {"estimated_h": t["estimated_h"], "real_h": 0.0,
                     "by_user": {}, "title": t["title"],
                     "sprint": t["sprint"], "area": t["area"],
                     "size": t["size"]}
               for tid, t in tasks.items()}
    by_user = {}

    for entry in entries:
        # duration in ISO 8601 like PT1H30M
        dur_str = entry.get("timeInterval", {}).get("duration", "PT0S")
        hours = parse_duration(dur_str)
        user_id   = entry.get("userId", "")
        user_name = users_map.get(user_id, user_id)
        tag_ids   = entry.get("tagIds", [])
        task_id   = next((tags_map[tid] for tid in tag_ids if tags_map.get(tid) in by_task), None)

        if not task_id:
            continue  # no matching task tag

        by_task[task_id]["real_h"] += hours
        by_task[task_id]["by_user"].setdefault(user_name, 0.0)
        by_task[task_id]["by_user"][user_name] += hours

        by_user.setdefault(user_name, {"total_h": 0.0, "by_task": {}})
        by_user[user_name]["total_h"]          += hours
        by_user[user_name]["by_task"].setdefault(task_id, 0.0)
        by_user[user_name]["by_task"][task_id] += hours

    return by_task, by_user


def parse_duration(dur: str) -> float:
    """Parse ISO 8601 duration string to hours."""
    import re
    if not dur or dur == "PT0S":
        return 0.0
    h = float(re.search(r"(\d+)H", dur).group(1)) if re.search(r"\d+H", dur) else 0
    m = float(re.search(r"(\d+)M", dur).group(1)) if re.search(r"\d+M", dur) else 0
    s = float(re.search(r"(\d+)S", dur).group(1)) if re.search(r"\d+S", dur) else 0
    return h + m / 60 + s / 3600


# ── EXCEL BUILDER ──────────────────────────────────────────────
def col(n: int) -> str:
    return get_column_letter(n)

def hdr(ws, row: int, col_idx: int, value, width=None, color="1E1E2E", font_color="FFFFFF", size=11, wrap=False):
    c = ws.cell(row=row, column=col_idx, value=value)
    c.font      = Font(bold=True, color=font_color, size=size, name="Arial")
    c.fill      = PatternFill("solid", start_color=color)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=wrap)
    if width:
        ws.column_dimensions[col(col_idx)].width = width
    return c

def cell(ws, row, col_idx, value, bold=False, color=None, font_color="000000",
         align="left", num_fmt=None, wrap=False):
    c = ws.cell(row=row, column=col_idx, value=value)
    c.font      = Font(bold=bold, color=font_color, size=10, name="Arial")
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    if color:
        c.fill  = PatternFill("solid", start_color=color)
    if num_fmt:
        c.number_format = num_fmt
    return c

def border_row(ws, row: int, cols: int):
    thin = Side(style="thin", color="E2E8F0")
    for c_idx in range(1, cols + 1):
        ws.cell(row=row, column=c_idx).border = Border(bottom=thin)

def status_color(pct: float) -> str:
    if pct >= 100:  return COLORS["danger"]
    if pct >= 80:   return COLORS["warn"]
    return COLORS["ok"]

def add_progress_cell(ws, row, col_idx, pct: float):
    """Fake progress bar via cell comment and colored background."""
    color = status_color(pct)
    c = ws.cell(row=row, column=col_idx, value=f"{pct:.0f}%")
    c.font      = Font(bold=True, color="FFFFFF", size=10, name="Arial")
    c.fill      = PatternFill("solid", start_color=color)
    c.alignment = Alignment(horizontal="center", vertical="center")


def build_excel(by_task: dict, by_user: dict, sprint_filter: int | None,
                output_path: str = "nexus_clockify_report.xlsx"):

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default sheet

    sprint_color = {1: COLORS["s1"], 2: COLORS["s2"], 3: COLORS["s3"]}

    # ── HOJA 1: Resumen ──────────────────────────────────────────
    ws1 = wb.create_sheet("📊 Resumen")
    ws1.row_dimensions[1].height = 40
    ws1.freeze_panes = "A3"

    # Title
    ws1.merge_cells("A1:G1")
    title_cell = ws1["A1"]
    title_cell.value   = "NexUS — Informe de horas × Clockify"
    title_cell.font    = Font(bold=True, size=16, color="FFFFFF", name="Arial")
    title_cell.fill    = PatternFill("solid", start_color=COLORS["header_bg"])
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    ws1.merge_cells("A2:G2")
    sub = ws1["A2"]
    sub.value     = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}  ·  Sprint filter: {'Todos' if not sprint_filter else f'Sprint {sprint_filter}'}"
    sub.font      = Font(size=10, color="94A3B8", name="Arial")
    sub.fill      = PatternFill("solid", start_color="111113")
    sub.alignment = Alignment(horizontal="center", vertical="center")

    # KPI cards row
    sprints_to_show = [sprint_filter] if sprint_filter else [1, 2, 3]
    row = 4
    ws1["A4"].value = "Sprint"
    ws1["B4"].value = "Tareas con horas"
    ws1["C4"].value = "H. Estimadas"
    ws1["D4"].value = "H. Reales"
    ws1["E4"].value = "% Completado"
    ws1["F4"].value = "Tareas en alerta (>80%)"
    ws1["G4"].value = "Tareas excedidas (>100%)"
    for ci in range(1, 8):
        hdr(ws1, 4, ci, ws1.cell(4, ci).value, color="1E1E2E")

    ws1.column_dimensions["A"].width = 12
    ws1.column_dimensions["B"].width = 18
    ws1.column_dimensions["C"].width = 16
    ws1.column_dimensions["D"].width = 14
    ws1.column_dimensions["E"].width = 16
    ws1.column_dimensions["F"].width = 22
    ws1.column_dimensions["G"].width = 22

    for s in sprints_to_show:
        tasks_s = {k: v for k, v in by_task.items() if v["sprint"] == s}
        total_est  = sum(t["estimated_h"] for t in tasks_s.values())
        total_real = sum(t["real_h"]      for t in tasks_s.values())
        with_hours = sum(1 for t in tasks_s.values() if t["real_h"] > 0)
        pct        = (total_real / total_est * 100) if total_est else 0
        alerts     = sum(1 for t in tasks_s.values() if 0 < t["estimated_h"] and 80 <= (t["real_h"]/t["estimated_h"]*100) < 100)
        exceeded   = sum(1 for t in tasks_s.values() if 0 < t["estimated_h"] and t["real_h"] >= t["estimated_h"])
        row += 1
        sc = sprint_color.get(s, "818CF8")
        cell(ws1, row, 1, f"Sprint {s}", bold=True, color=sc, font_color="FFFFFF", align="center")
        cell(ws1, row, 2, f"{with_hours} / {len(tasks_s)}", align="center")
        cell(ws1, row, 3, total_est,  num_fmt="0.0", align="center")
        cell(ws1, row, 4, total_real, num_fmt="0.0", align="center")
        add_progress_cell(ws1, row, 5, pct)
        cell(ws1, row, 6, alerts,   color="FEF3C7" if alerts   else None, bold=bool(alerts),   align="center")
        cell(ws1, row, 7, exceeded, color="FEE2E2" if exceeded else None, bold=bool(exceeded), align="center")
        border_row(ws1, row, 7)

    # ── HOJA 2: Real vs Estimado por tarea ──────────────────────
    ws2 = wb.create_sheet("📋 Tarea vs Estimado")
    ws2.freeze_panes = "A3"
    ws2.row_dimensions[1].height = 30

    headers2 = ["ID", "Sprint", "Módulo", "Tarea", "Talla", "H. Estimadas",
                 "H. Reales", "Diferencia", "% Uso", "Estado"]
    widths2  = [10, 9, 20, 48, 8, 14, 12, 12, 10, 12]
    ws2.merge_cells(f"A1:{col(len(headers2))}1")
    ws2["A1"].value = "Horas Reales vs Estimadas por Tarea"
    ws2["A1"].font  = Font(bold=True, size=13, color="FFFFFF", name="Arial")
    ws2["A1"].fill  = PatternFill("solid", start_color=COLORS["header_bg"])
    ws2["A1"].alignment = Alignment(horizontal="center", vertical="center")

    for ci, (h, w) in enumerate(zip(headers2, widths2), 1):
        hdr(ws2, 2, ci, h, width=w)

    tasks_sorted = sorted(by_task.items(), key=lambda x: (x[1]["sprint"], x[1]["area"], x[0]))
    for r_idx, (tid, t) in enumerate(tasks_sorted, 3):
        if sprint_filter and t["sprint"] != sprint_filter:
            continue
        est  = t["estimated_h"]
        real = t["real_h"]
        diff = real - est
        pct  = (real / est * 100) if est else 0
        bg   = COLORS["light_row"] if r_idx % 2 == 0 else COLORS["white"]
        sc   = sprint_color.get(t["sprint"], "818CF8")

        cell(ws2, r_idx, 1,  tid,     bold=True, color=sc, font_color="FFFFFF", align="center")
        cell(ws2, r_idx, 2,  f"S{t['sprint']}", color=sc+"22", align="center")
        cell(ws2, r_idx, 3,  t["area"],  align="left")
        cell(ws2, r_idx, 4,  t["title"], align="left", wrap=True)
        cell(ws2, r_idx, 5,  t["size"],  align="center")
        cell(ws2, r_idx, 6,  est,  num_fmt="0.0", align="center")
        cell(ws2, r_idx, 7,  real, num_fmt="0.0", align="center",
             color="DCF8E6" if real > 0 else None)
        diff_color = "FEE2E2" if diff > 0 else ("DCF8E6" if diff < 0 else None)
        cell(ws2, r_idx, 8,  diff, num_fmt="+0.0;-0.0;0.0", align="center",
             color=diff_color, font_color="EF4444" if diff > 0 else "22C55E")
        add_progress_cell(ws2, r_idx, 9, pct)

        if real == 0:        estado = "⬜ Sin iniciar"
        elif pct >= 100:     estado = "🔴 Excedida"
        elif pct >= 80:      estado = "🟡 En riesgo"
        else:                estado = "🟢 En curso"
        cell(ws2, r_idx, 10, estado, align="center")
        ws2.row_dimensions[r_idx].height = 22
        border_row(ws2, r_idx, len(headers2))

    ws2.auto_filter.ref = f"A2:{col(len(headers2))}2"

    # ── HOJA 3: Horas por persona ────────────────────────────────
    ws3 = wb.create_sheet("👥 Horas por Persona")
    ws3.freeze_panes = "B3"

    # Get all task IDs with hours
    active_tasks = [tid for tid, t in by_task.items() if t["real_h"] > 0]
    users_sorted = sorted(by_user.keys())

    ws3["A1"].value = "Horas registradas por persona y tarea"
    ws3["A1"].font  = Font(bold=True, size=13, color="FFFFFF", name="Arial")
    ws3["A1"].fill  = PatternFill("solid", start_color=COLORS["header_bg"])
    ws3.merge_cells(f"A1:{col(len(users_sorted)+2)}1")
    ws3["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws3.row_dimensions[1].height = 30

    hdr(ws3, 2, 1, "Tarea ID",   width=12)
    hdr(ws3, 2, 2, "Módulo",     width=22)
    for ci, u in enumerate(users_sorted, 3):
        hdr(ws3, 2, ci, u, width=14)
    hdr(ws3, 2, len(users_sorted)+3, "TOTAL", width=12)

    for r_idx, tid in enumerate(active_tasks, 3):
        t = by_task[tid]
        sc = sprint_color.get(t["sprint"], "818CF8")
        cell(ws3, r_idx, 1, tid,      bold=True, color=sc, font_color="FFFFFF", align="center")
        cell(ws3, r_idx, 2, t["area"], align="left")
        row_total = 0
        for ci, u in enumerate(users_sorted, 3):
            h = t["by_user"].get(u, 0)
            row_total += h
            cell(ws3, r_idx, ci, h if h else None, num_fmt="0.0", align="center",
                 color="DCF8E6" if h > 0 else None)
        cell(ws3, r_idx, len(users_sorted)+3, row_total,
             num_fmt="0.0", align="center", bold=True)
        border_row(ws3, r_idx, len(users_sorted)+3)
        ws3.row_dimensions[r_idx].height = 20

    # Total row
    total_row = len(active_tasks) + 3
    hdr(ws3, total_row, 1, "TOTAL", color="1E1E2E")
    hdr(ws3, total_row, 2, "", color="1E1E2E")
    for ci, u in enumerate(users_sorted, 3):
        total_u = by_user[u]["total_h"]
        hdr(ws3, total_row, ci, f"{total_u:.1f}h", color="1E1E2E")
    grand = sum(t["total_h"] for t in by_user.values())
    hdr(ws3, total_row, len(users_sorted)+3, f"{grand:.1f}h", color="1E1E2E")

    # ── HOJA 4: Alertas ─────────────────────────────────────────
    ws4 = wb.create_sheet("⚠️ Alertas")
    ws4.freeze_panes = "A3"

    ws4["A1"].value = "Tareas en riesgo o excedidas"
    ws4["A1"].font  = Font(bold=True, size=13, color="FFFFFF", name="Arial")
    ws4["A1"].fill  = PatternFill("solid", start_color="7F1D1D")
    ws4.merge_cells("A1:H1")
    ws4["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws4.row_dimensions[1].height = 30

    headers4 = ["ID", "Sprint", "Módulo", "Tarea", "H. Est.", "H. Real", "% Uso", "Acción recomendada"]
    widths4  = [10, 9, 20, 44, 10, 10, 10, 38]
    for ci, (h, w) in enumerate(zip(headers4, widths4), 1):
        hdr(ws4, 2, ci, h, width=w, color="7F1D1D")

    alerts = [(tid, t) for tid, t in by_task.items()
              if t["estimated_h"] > 0 and t["real_h"] > 0
              and (t["real_h"] / t["estimated_h"]) >= 0.8]
    alerts.sort(key=lambda x: x[1]["real_h"]/x[1]["estimated_h"], reverse=True)

    for r_idx, (tid, t) in enumerate(alerts, 3):
        pct  = t["real_h"] / t["estimated_h"] * 100
        is_exc = pct >= 100
        bg   = "FEE2E2" if is_exc else "FEF3C7"
        accion = ("⛔ Renegociar alcance o dividir tarea" if is_exc
                  else "⚠️ Revisar progreso y re-estimar si necesario")
        cell(ws4, r_idx, 1, tid,           bold=True, color=bg, align="center")
        cell(ws4, r_idx, 2, f"S{t['sprint']}",          color=bg, align="center")
        cell(ws4, r_idx, 3, t["area"],                   color=bg)
        cell(ws4, r_idx, 4, t["title"],                  color=bg, wrap=True)
        cell(ws4, r_idx, 5, t["estimated_h"], num_fmt="0.0", color=bg, align="center")
        cell(ws4, r_idx, 6, t["real_h"],      num_fmt="0.0", color=bg, align="center")
        add_progress_cell(ws4, r_idx, 7, pct)
        cell(ws4, r_idx, 8, accion,                      color=bg, wrap=True)
        ws4.row_dimensions[r_idx].height = 28
        border_row(ws4, r_idx, 8)

    if not alerts:
        ws4["A3"].value = "✅ No hay tareas en riesgo ni excedidas."
        ws4["A3"].font  = Font(size=12, color="22C55E", name="Arial")

    # Tab colors
    ws1.sheet_properties.tabColor = "1E1E2E"
    ws2.sheet_properties.tabColor = COLORS["s1"]
    ws3.sheet_properties.tabColor = COLORS["s2"]
    ws4.sheet_properties.tabColor = COLORS["danger"]

    wb.save(output_path)
    print(f"✅ Informe generado: {output_path}")
    print(f"   Tareas procesadas: {len(by_task)}")
    print(f"   Personas con horas: {len(by_user)}")
    print(f"   Alertas detectadas: {len(alerts)}")


# ── MAIN ───────────────────────────────────────────────────────
def main():
    ap = argparse.ArgumentParser(description="NexUS × Clockify Report")
    ap.add_argument("--api-key",   required=True, help="Clockify API Key (perfil → API Key)")
    ap.add_argument("--workspace", default=None,  help="Workspace ID (opcional, usa el primero)")
    ap.add_argument("--sprint",    type=int, default=None, help="Filtrar por sprint (1, 2 o 3)")
    ap.add_argument("--backlog",   default="nexus-backlog.json", help="Ruta al JSON del backlog")
    ap.add_argument("--output",    default="nexus_clockify_report.xlsx", help="Ruta de salida")
    ap.add_argument("--start",     default=None, help="Fecha inicio ISO (ej: 2026-02-19T00:00:00Z)")
    ap.add_argument("--end",       default=None, help="Fecha fin ISO")
    args = ap.parse_args()

    print("🔗 Conectando con Clockify...")
    client = ClockifyClient(args.api_key)

    ws_id = args.workspace or client.workspace()
    print(f"   Workspace: {ws_id}")

    print("👥 Cargando usuarios...")
    users_map = client.users(ws_id)
    print(f"   {len(users_map)} usuarios encontrados")

    print("🏷️  Cargando tags...")
    tags_map = client.tags(ws_id)
    print(f"   {len(tags_map)} tags encontrados")

    print("⏱️  Descargando time entries...")
    entries = client.entries(ws_id, args.start, args.end)
    print(f"   {len(entries)} entradas descargadas")

    print("📦 Cargando backlog...")
    tasks = load_backlog(args.backlog)
    print(f"   {len(tasks)} tareas en el backlog")

    print("🔀 Cruzando datos...")
    by_task, by_user = process_entries(entries, tags_map, users_map, tasks)

    print("📊 Generando Excel...")
    build_excel(by_task, by_user, args.sprint, args.output)


if __name__ == "__main__":
    main()
